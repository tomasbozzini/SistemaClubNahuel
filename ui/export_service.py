# ui/export_service.py
"""
Exportación de datos a Excel (.xlsx) y PDF (.pdf).
Usa openpyxl para Excel y fpdf2 para PDF.
"""
import os
from datetime import datetime
from tkinter import filedialog, messagebox


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _pedir_ruta(titulo: str, ext: str, defecto: str) -> str | None:
    ruta = filedialog.asksaveasfilename(
        title=titulo,
        defaultextension=f".{ext}",
        initialfile=f"{defecto}_{_ts()}.{ext}",
        filetypes=[(ext.upper(), f"*.{ext}"), ("Todos", "*.*")],
    )
    return ruta or None


# ── Excel ─────────────────────────────────────────────────────────────────────

def _excel_style(ws, header_fill, cols_width):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, width in enumerate(cols_width, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            cell.font      = Font(size=10)

    ws.row_dimensions[1].height = 22


def exportar_excel_reservas(filas: list) -> None:
    """
    filas: [(id, cliente, cancha, tipo, fecha, hora, notas, telefono), ...]
    """
    if not filas:
        messagebox.showwarning("Sin datos", "No hay reservas para exportar.")
        return

    ruta = _pedir_ruta("Guardar Excel de reservas", "xlsx", "reservas")
    if not ruta:
        return

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"

        headers = ["ID", "Cliente", "Celular", "Cancha", "Tipo", "Fecha", "Hora", "Notas"]
        ws.append(headers)

        for f in filas:
            # f: (id[0], cliente[1], cancha[2], tipo[3], fecha[4], hora[5], notas[6], telefono[7])
            ws.append([f[0], f[1], f[7], f[2], f[3], f[4], f[5], f[6]])

        _excel_style(ws, "1A6B2A", [6, 22, 16, 18, 10, 12, 8, 30])
        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{os.path.basename(ruta)}")
    except Exception as e:
        messagebox.showerror("Error al exportar", str(e))


def exportar_excel_financiero(filas: list) -> None:
    """
    filas: [(id, cliente, cancha, tipo, fecha, hora_ini, hora_fin, dur_min, estado, precio), ...]
    """
    if not filas:
        messagebox.showwarning("Sin datos", "No hay registros para exportar.")
        return

    ruta = _pedir_ruta("Guardar Excel financiero", "xlsx", "financiero")
    if not ruta:
        return

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Financiero"

        headers = ["ID", "Cliente", "Cancha", "Tipo", "Fecha", "Inicio", "Fin", "Duración (min)", "Estado", "Precio"]
        ws.append(headers)

        for f in filas:
            dur = f[7]
            dur_str = f"{dur // 60}h {dur % 60:02d}m" if dur % 60 else f"{dur // 60}h"
            precio = f[9] if f[8] == "completada" else 0
            ws.append([f[0], f[1], f[2], f[3], f[4], f[5], f[6], dur_str, f[8], precio])

        _excel_style(ws, "7A5800", [6, 22, 18, 10, 12, 8, 8, 14, 14, 14])

        # Resaltar total en la última fila
        from openpyxl.styles import Font
        total = sum(f[9] for f in filas if f[8] == "completada")
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "", "TOTAL COBRADO", total])
        last_row = ws.max_row
        ws.cell(last_row, 9).font = Font(bold=True, size=10)
        ws.cell(last_row, 10).font = Font(bold=True, size=10)

        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{os.path.basename(ruta)}")
    except Exception as e:
        messagebox.showerror("Error al exportar", str(e))


# ── PDF ───────────────────────────────────────────────────────────────────────

def _fmt_peso(valor) -> str:
    try:
        return f"$ {int(float(valor)):,}".replace(",", ".")
    except Exception:
        return "$ 0"


def _pdf_safe(text) -> str:
    """Convierte texto a latin-1 seguro para las fuentes built-in de fpdf2."""
    _repl = {
        "\u2014": "-", "\u2013": "-",   # em/en dash
        "\u2192": "->", "\u2190": "<-",
        "\u21BA": "~",  "\u2B07": "v",
        "\u25C9": "*",  "\u2726": "*",
        "\u25CE": "*",  "\u25C8": "*",
    }
    result = []
    for c in str(text):
        c = _repl.get(c, c)
        result.append(c if ord(c) < 256 else "?")
    return "".join(result)


def exportar_pdf_reservas(filas: list) -> None:
    if not filas:
        messagebox.showwarning("Sin datos", "No hay reservas para exportar.")
        return

    ruta = _pedir_ruta("Guardar PDF de reservas", "pdf", "reservas")
    if not ruta:
        return

    try:
        from fpdf import FPDF

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=12)

        # Título
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(30, 30, 30)
        pdf.cell(0, 10, "Club Nahuel - Listado de Reservas", ln=True, align="C")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 6, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
        pdf.ln(4)

        # Tabla
        headers = ["ID", "Cliente", "Celular", "Cancha", "Tipo", "Fecha", "Hora", "Notas"]
        col_w   = [12,   52,        36,        38,       22,     28,      18,     60]

        pdf.set_fill_color(26, 107, 42)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        for h, w in zip(headers, col_w):
            pdf.cell(w, 8, h, border=1, fill=True, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        for i, f in enumerate(filas):
            fill = i % 2 == 0
            pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(30, 30, 30)
            row = [str(f[0]), f[1], f[7], f[2], f[3], f[4], f[5], f[6]]
            for val, w in zip(row, col_w):
                pdf.cell(w, 7, _pdf_safe(val)[:30], border=1, fill=fill, align="C")
            pdf.ln()

        pdf.output(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{os.path.basename(ruta)}")
    except Exception as e:
        messagebox.showerror("Error al exportar", str(e))


def exportar_pdf_financiero(filas: list) -> None:
    if not filas:
        messagebox.showwarning("Sin datos", "No hay registros para exportar.")
        return

    ruta = _pedir_ruta("Guardar PDF financiero", "pdf", "financiero")
    if not ruta:
        return

    try:
        from fpdf import FPDF

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=12)

        # Título
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(30, 30, 30)
        pdf.cell(0, 10, "Club Nahuel - Historial Financiero", ln=True, align="C")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 6, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
        pdf.ln(4)

        # Tabla
        headers = ["ID", "Cliente", "Cancha", "Tipo", "Fecha", "Inicio", "Fin", "Duración", "Estado", "Precio"]
        col_w   = [12,   50,        38,       22,     28,      16,       16,    20,          26,       26]

        pdf.set_fill_color(122, 88, 0)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        for h, w in zip(headers, col_w):
            pdf.cell(w, 8, h, border=1, fill=True, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        total = 0.0
        for i, f in enumerate(filas):
            fill = i % 2 == 0
            pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(30, 30, 30)
            dur = f[7]
            dur_str = f"{dur // 60}h {dur % 60:02d}m" if dur % 60 else f"{dur // 60}h"
            precio_val = f[9] if f[8] == "completada" else 0.0
            total += precio_val
            precio_str = _fmt_peso(precio_val) if f[8] == "completada" else "-"
            row = [str(f[0]), f[1], f[2], f[3], f[4], f[5], f[6], dur_str, f[8], precio_str]
            for val, w in zip(row, col_w):
                pdf.cell(w, 7, _pdf_safe(val)[:28], border=1, fill=fill, align="C")
            pdf.ln()

        # Total
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(30, 30, 30)
        total_w = sum(col_w[:-1])
        pdf.cell(total_w, 8, "TOTAL COBRADO", border=1, align="R")
        pdf.cell(col_w[-1], 8, _fmt_peso(total), border=1, align="C")
        pdf.ln()

        pdf.output(ruta)
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{os.path.basename(ruta)}")
    except Exception as e:
        messagebox.showerror("Error al exportar", str(e))
