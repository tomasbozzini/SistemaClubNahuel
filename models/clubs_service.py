# models/clubs_service.py
# Servicio para gestión de clubes (solo superadmin).

from sqlalchemy import text
from db.database import engine, get_connection
from models.club import Club


# ── Lectura ───────────────────────────────────────────────────────────────────

def listar_todos_los_clubs() -> list[dict]:
    """Retorna todos los clubes con todos sus campos."""
    with engine.connect() as conn:
        filas = conn.execute(text("""
            SELECT id, nombre, ciudad, plan, activo, modo_mantenimiento,
                   fecha_instalacion, monto_implementacion, precio_mensual,
                   dia_vencimiento, fecha_ultimo_pago, estado_pago, notas
            FROM clubs
            ORDER BY id
        """)).fetchall()
    return [dict(f._mapping) for f in filas]


def get_club(club_id: int) -> dict | None:
    with engine.connect() as conn:
        fila = conn.execute(
            text("SELECT * FROM clubs WHERE id = :id"),
            {"id": club_id}
        ).fetchone()
    return dict(fila._mapping) if fila else None


def get_metricas_dashboard() -> dict:
    """Retorna todas las métricas del dashboard en una sola llamada."""
    with engine.connect() as conn:
        r = conn.execute(text("""
            SELECT
                COUNT(*) FILTER (WHERE activo = TRUE)  AS clubes_activos,
                COUNT(*) FILTER (WHERE activo = FALSE) AS clubes_inactivos,
                COALESCE(SUM(precio_mensual) FILTER (WHERE activo = TRUE), 0)
                    AS ingreso_proyectado,
                COUNT(*) FILTER (
                    WHERE activo = TRUE AND estado_pago = 'vencido'
                ) AS pagos_vencidos,
                COUNT(*) FILTER (
                    WHERE activo = TRUE
                    AND estado_pago != 'vencido'
                    AND dia_vencimiento BETWEEN
                        EXTRACT(DAY FROM CURRENT_DATE)
                        AND EXTRACT(DAY FROM CURRENT_DATE + INTERVAL '7 days')
                ) AS vencen_pronto
            FROM clubs
        """)).fetchone()

        total_usuarios = conn.execute(text(
            "SELECT COUNT(*) FROM usuarios WHERE rol != 'superadmin' AND activo = TRUE"
        )).scalar() or 0

        reservas_hoy = conn.execute(text(
            "SELECT COUNT(*) FROM reservas "
            "WHERE fecha = CURRENT_DATE AND estado = 'confirmada'"
        )).scalar() or 0

        pagado_mes = conn.execute(text("""
            SELECT COALESCE(SUM(monto), 0) FROM pagos_mantenimiento
            WHERE EXTRACT(MONTH FROM fecha_pago) = EXTRACT(MONTH FROM CURRENT_DATE)
              AND EXTRACT(YEAR  FROM fecha_pago) = EXTRACT(YEAR  FROM CURRENT_DATE)
        """)).scalar() or 0

    return {
        "clubes_activos":     int(r.clubes_activos),
        "clubes_inactivos":   int(r.clubes_inactivos),
        "ingreso_proyectado": float(r.ingreso_proyectado),
        "pagos_vencidos":     int(r.pagos_vencidos),
        "vencen_pronto":      int(r.vencen_pronto),
        "total_usuarios":     int(total_usuarios),
        "reservas_hoy":       int(reservas_hoy),
        "pagado_mes":         float(pagado_mes),
    }


def listar_pagos(club_id: int = None, limite: int = 50) -> list[dict]:
    """Retorna historial de pagos, opcionalmente filtrado por club."""
    where = "WHERE p.club_id = :club_id" if club_id else ""
    params = {"limite": limite}
    if club_id:
        params["club_id"] = club_id
    with engine.connect() as conn:
        filas = conn.execute(text(f"""
            SELECT p.id, p.club_id, c.nombre AS club_nombre,
                   p.fecha_pago, p.monto, p.periodo, p.notas, p.creado_en
            FROM pagos_mantenimiento p
            JOIN clubs c ON p.club_id = c.id
            {where}
            ORDER BY p.fecha_pago DESC, p.creado_en DESC
            LIMIT :limite
        """), params).fetchall()
    return [dict(f._mapping) for f in filas]


def listar_logs_todos(
    club_id: int = None,
    accion: str = None,
    limite: int = 100,
) -> list[dict]:
    """Retorna logs de acceso de todos los clubes con filtros opcionales."""
    filtros, params = [], {"limite": limite}
    if club_id:
        filtros.append("l.club_id = :club_id")
        params["club_id"] = club_id
    if accion:
        filtros.append("l.accion = :accion")
        params["accion"] = accion
    where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
    with engine.connect() as conn:
        filas = conn.execute(text(f"""
            SELECT l.id, l.username, l.accion, l.detalle, l.hostname,
                   l.timestamp, l.club_id, c.nombre AS club_nombre
            FROM logs_acceso l
            LEFT JOIN clubs c ON l.club_id = c.id
            {where}
            ORDER BY l.timestamp DESC
            LIMIT :limite
        """), params).fetchall()
    return [dict(f._mapping) for f in filas]


def club_en_mantenimiento(club_id: int) -> bool:
    """Retorna True si el club está en modo mantenimiento."""
    if not club_id:
        return False
    with engine.connect() as conn:
        val = conn.execute(
            text("SELECT modo_mantenimiento FROM clubs WHERE id = :id"),
            {"id": club_id}
        ).scalar()
    return bool(val)


# ── Escritura ─────────────────────────────────────────────────────────────────

def crear_club(
    nombre: str,
    ciudad: str,
    plan: str,
    monto_implementacion: float,
    precio_mensual: float,
    dia_vencimiento: int,
    notas: str = "",
) -> int:
    """Crea un nuevo club. Retorna el id generado."""
    with engine.connect() as conn:
        result = conn.execute(text("""
            INSERT INTO clubs
                (nombre, ciudad, plan, monto_implementacion, precio_mensual,
                 dia_vencimiento, estado_pago, notas)
            VALUES
                (:nombre, :ciudad, :plan, :monto, :mensual,
                 :dia_venc, 'al_dia', :notas)
            RETURNING id
        """), {
            "nombre": nombre.strip(),
            "ciudad": ciudad.strip(),
            "plan":   plan,
            "monto":  monto_implementacion,
            "mensual": precio_mensual,
            "dia_venc": dia_vencimiento,
            "notas": notas.strip() or None,
        })
        nuevo_id = result.scalar()
        conn.commit()
    return nuevo_id


def actualizar_club(
    club_id: int,
    nombre: str,
    ciudad: str,
    plan: str,
    monto_implementacion: float,
    precio_mensual: float,
    dia_vencimiento: int,
    estado_pago: str,
    notas: str = "",
):
    with engine.connect() as conn:
        conn.execute(text("""
            UPDATE clubs SET
                nombre               = :nombre,
                ciudad               = :ciudad,
                plan                 = :plan,
                monto_implementacion = :monto,
                precio_mensual       = :mensual,
                dia_vencimiento      = :dia_venc,
                estado_pago          = :estado_pago,
                notas                = :notas
            WHERE id = :id
        """), {
            "nombre":      nombre.strip(),
            "ciudad":      ciudad.strip(),
            "plan":        plan,
            "monto":       monto_implementacion,
            "mensual":     precio_mensual,
            "dia_venc":    dia_vencimiento,
            "estado_pago": estado_pago,
            "notas":       notas.strip() or None,
            "id":          club_id,
        })
        conn.commit()


def toggle_activo(club_id: int) -> bool:
    """Activa/desactiva un club. Retorna el nuevo estado."""
    with engine.connect() as conn:
        nuevo = conn.execute(text("""
            UPDATE clubs SET activo = NOT activo
            WHERE id = :id
            RETURNING activo
        """), {"id": club_id}).scalar()
        conn.commit()
    return bool(nuevo)


def toggle_mantenimiento(club_id: int) -> bool:
    """Activa/desactiva modo mantenimiento. Retorna el nuevo estado."""
    with engine.connect() as conn:
        nuevo = conn.execute(text("""
            UPDATE clubs SET modo_mantenimiento = NOT modo_mantenimiento
            WHERE id = :id
            RETURNING modo_mantenimiento
        """), {"id": club_id}).scalar()
        conn.commit()
    return bool(nuevo)


def registrar_pago(
    club_id: int,
    monto: float,
    periodo: str = "",
    notas: str = "",
):
    """Registra un pago mensual y actualiza estado_pago y fecha_ultimo_pago del club."""
    with engine.connect() as conn:
        conn.execute(text("""
            INSERT INTO pagos_mantenimiento (club_id, monto, periodo, notas)
            VALUES (:club_id, :monto, :periodo, :notas)
        """), {
            "club_id": club_id,
            "monto":   monto,
            "periodo": periodo.strip() or None,
            "notas":   notas.strip() or None,
        })
        conn.execute(text("""
            UPDATE clubs
            SET fecha_ultimo_pago = CURRENT_DATE,
                estado_pago       = 'al_dia'
            WHERE id = :id
        """), {"id": club_id})
        conn.commit()


def exportar_datos_club(club_id: int, ruta_destino: str):
    """Exporta reservas y clientes del club a un archivo Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Hoja de reservas
        ws_r = wb.active
        ws_r.title = "Reservas"
        with engine.connect() as conn:
            club = conn.execute(
                text("SELECT nombre FROM clubs WHERE id = :id"), {"id": club_id}
            ).fetchone()
            club_nombre = club.nombre if club else f"Club {club_id}"

            reservas = conn.execute(text("""
                SELECT r.id, r.nombre_cliente, r.telefono_cliente,
                       c.nombre AS cancha, r.fecha, r.hora_inicio, r.hora_fin,
                       r.estado, r.estado_pago, r.precio_total, r.notas
                FROM reservas r
                JOIN canchas c ON r.cancha_id = c.id
                WHERE r.club_id = :cid
                ORDER BY r.fecha DESC, r.hora_inicio DESC
            """), {"cid": club_id}).fetchall()

            clientes = conn.execute(text("""
                SELECT id, nombre, telefono, email, creado_en
                FROM clientes WHERE club_id = :cid ORDER BY nombre
            """), {"cid": club_id}).fetchall()

        # Encabezados reservas
        hdrs_r = ["ID", "Cliente", "Teléfono", "Cancha", "Fecha",
                  "Inicio", "Fin", "Estado", "Pago", "Precio", "Notas"]
        for col, h in enumerate(hdrs_r, 1):
            c = ws_r.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1A1A1A")
        for row_i, r in enumerate(reservas, 2):
            for col_i, val in enumerate(r, 1):
                ws_r.cell(row=row_i, column=col_i, value=str(val) if val is not None else "")

        # Hoja de clientes
        ws_c = wb.create_sheet("Clientes")
        hdrs_c = ["ID", "Nombre", "Teléfono", "Email", "Registrado"]
        for col, h in enumerate(hdrs_c, 1):
            c = ws_c.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1A1A1A")
        for row_i, r in enumerate(clientes, 2):
            for col_i, val in enumerate(r, 1):
                ws_c.cell(row=row_i, column=col_i, value=str(val) if val is not None else "")

        wb.save(ruta_destino)
        return True, club_nombre
    except Exception as e:
        return False, str(e)
